"""OASIS MMSE metadata loading utilities."""

from __future__ import annotations

from pathlib import Path

from brainage.data.hcp_mmse import HCPMMSEExample

try:
    from openpyxl import load_workbook
except ModuleNotFoundError as exc:  # pragma: no cover - dependency guard
    load_workbook = None
    _IMPORT_ERROR = exc
else:
    _IMPORT_ERROR = None


def require_oasis_dependencies() -> None:
    if _IMPORT_ERROR is not None:
        raise RuntimeError(
            "OASIS MMSE loading requires `openpyxl`. Install it with `pip install -r requirements.txt`."
        ) from _IMPORT_ERROR


def discover_oasis_mmse_examples(
    metadata_path: Path,
    image_dir: Path,
    subject_id_column: str = "subject_id",
    target_column: str = "mmse",
    age_column: str = "age",
    sex_column: str = "gender",
) -> list[HCPMMSEExample]:
    require_oasis_dependencies()

    image_lookup = {path.name: path for path in sorted(image_dir.glob('OAS*.nii'))}
    workbook = load_workbook(metadata_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"OASIS metadata workbook is empty: {metadata_path}")

    header = [str(value).strip() if value is not None else "" for value in rows[0]]
    column_index = {name: idx for idx, name in enumerate(header)}
    required = [subject_id_column, target_column, age_column, sex_column]
    missing = [name for name in required if name not in column_index]
    if missing:
        raise KeyError(f"Missing required OASIS metadata columns: {missing}")

    examples: list[HCPMMSEExample] = []
    seen_subject_ids: set[str] = set()
    for row_index, row in enumerate(rows[1:], start=2):
        subject_id = str(row[column_index[subject_id_column]] or "").strip()
        mmse_value = row[column_index[target_column]]
        age_value = row[column_index[age_column]]
        sex_value = str(row[column_index[sex_column]] or "").strip().upper()

        if not subject_id or mmse_value is None:
            continue
        if subject_id in seen_subject_ids:
            continue

        image_path = image_lookup.get(subject_id)
        if image_path is None:
            continue

        examples.append(
            HCPMMSEExample(
                subject_id=subject_id,
                image_path=image_path,
                mmse=float(mmse_value),
                age=float(age_value) if age_value is not None else None,
                sex=sex_value or None,
            )
        )
        seen_subject_ids.add(subject_id)

    if not examples:
        raise ValueError("No OASIS examples were matched between the Excel metadata and MRI files.")

    return examples
