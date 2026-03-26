from __future__ import annotations

from collections import defaultdict
from typing import Iterable
from dataclasses import dataclass
from pathlib import Path
import random

from brainage.data.hcp_mmse import HCPMMSEDataset, HCPMMSEExample, discover_hcp_mmse_examples
from brainage.data.oasis_mmse import require_oasis_dependencies

try:
    from openpyxl import load_workbook
except ModuleNotFoundError:
    load_workbook = None




def _resolve_metadata_column(column_index: dict[str, int], desired_name: str) -> str:
    normalized = {name.strip().lower(): name for name in column_index}
    key = desired_name.strip().lower()
    if key not in normalized:
        raise KeyError(f"Column '{desired_name}' was not found in OASIS headers: {sorted(column_index)}")
    return normalized[key]


def _resolve_first_matching_column(column_index: dict[str, int], candidates: Iterable[str] | None) -> str | None:
    if candidates is None:
        return None
    normalized = {name.strip().lower(): name for name in column_index}
    for candidate in candidates:
        key = str(candidate).strip().lower()
        if key in normalized:
            return normalized[key]
    return None


def _normalize_filter_value(value) -> str:
    text = str(value or "").strip().lower()
    if text in {"0", "0.0", "cdr 0", "cdr=0"}:
        return "normal"
    return text


def _default_oasis_filter_column_candidates() -> list[str]:
    return [
        "diagnosis",
        "dx",
        "group",
        "cdr",
        "cdr_global",
        "clinical dementia rating",
        "cdr-sb",
    ]

@dataclass(frozen=True)
class MultiSourceMMSEExample:
    subject_id: str
    image_path: Path
    mmse: float
    age: float | None
    sex: str | None
    domain_name: str
    domain_index: int


class MultiSourceMMSEDataset(HCPMMSEDataset):
    def __init__(self, examples: list[MultiSourceMMSEExample], *args, **kwargs) -> None:
        super().__init__(examples=examples, *args, **kwargs)

    def __getitem__(self, index: int) -> dict[str, object]:
        item = super().__getitem__(index)
        example = self.examples[index]
        item["domain_name"] = example.domain_name
        item["domain_index"] = example.domain_index
        return item


def _convert_examples(examples: list[HCPMMSEExample], domain_name: str, domain_index: int) -> list[MultiSourceMMSEExample]:
    return [
        MultiSourceMMSEExample(
            subject_id=example.subject_id,
            image_path=example.image_path,
            mmse=example.mmse,
            age=example.age,
            sex=example.sex,
            domain_name=domain_name,
            domain_index=domain_index,
        )
        for example in examples
    ]


def discover_hcp_source_examples(
    csv_path: Path,
    image_dir: Path,
    subject_id_column: str = "subject_id",
    target_column: str = "mmse",
    age_column: str | None = "age",
    sex_column: str | None = "sex",
    domain_name: str = "hcp",
    domain_index: int = 0,
) -> list[MultiSourceMMSEExample]:
    examples = discover_hcp_mmse_examples(
        csv_path=csv_path,
        image_dir=image_dir,
        subject_id_column=subject_id_column,
        target_column=target_column,
        age_column=age_column,
        sex_column=sex_column,
    )
    return _convert_examples(examples, domain_name=domain_name, domain_index=domain_index)


def discover_oasis_source_examples(
    metadata_path: Path,
    image_dir: Path,
    subject_id_column: str = "subject_id",
    target_column: str = "mmse",
    age_column: str = "age",
    sex_column: str = "gender",
    filter_column: str | None = None,
    filter_values: list[str] | None = None,
    domain_name: str = "oasis_nc",
    domain_index: int = 1,
) -> list[MultiSourceMMSEExample]:
    require_oasis_dependencies()
    if load_workbook is None:
        raise RuntimeError("openpyxl is required for OASIS metadata loading")

    image_lookup = {path.name: path for path in sorted(image_dir.glob("OAS*.nii"))}
    workbook = load_workbook(metadata_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"OASIS metadata workbook is empty: {metadata_path}")

    header = [str(value).strip() if value is not None else "" for value in rows[0]]
    column_index = {name: idx for idx, name in enumerate(header)}
    subject_id_column = _resolve_metadata_column(column_index, subject_id_column)
    target_column = _resolve_metadata_column(column_index, target_column)
    age_column = _resolve_metadata_column(column_index, age_column)
    sex_column = _resolve_metadata_column(column_index, sex_column)

    filter_column_name = None
    filter_candidates = []
    if filter_column:
        filter_candidates.append(filter_column)
    filter_candidates.extend(_default_oasis_filter_column_candidates())
    filter_column_name = _resolve_first_matching_column(column_index, filter_candidates)

    normalized_filter_values = None
    if filter_values:
        normalized_filter_values = {_normalize_filter_value(value) for value in filter_values}
        normalized_filter_values.update({"normal", "control", "cn", "nc", "cognitively normal"})

    examples: list[MultiSourceMMSEExample] = []
    seen_subject_ids: set[str] = set()
    for row in rows[1:]:
        subject_id = str(row[column_index[subject_id_column]] or "").strip()
        mmse_value = row[column_index[target_column]]
        age_value = row[column_index[age_column]]
        sex_value = str(row[column_index[sex_column]] or "").strip().upper()

        if normalized_filter_values is not None:
            if filter_column_name is None:
                raise KeyError(f"Missing OASIS filter column. Available headers: {sorted(column_index)}")
            filter_value = _normalize_filter_value(row[column_index[filter_column_name]])
            if filter_value not in normalized_filter_values:
                continue

        if not subject_id or mmse_value is None or subject_id in seen_subject_ids:
            continue

        image_path = image_lookup.get(subject_id)
        if image_path is None:
            continue

        examples.append(
            MultiSourceMMSEExample(
                subject_id=subject_id,
                image_path=image_path,
                mmse=float(mmse_value),
                age=float(age_value) if age_value is not None else None,
                sex=sex_value or None,
                domain_name=domain_name,
                domain_index=domain_index,
            )
        )
        seen_subject_ids.add(subject_id)

    if not examples:
        raise ValueError("No OASIS source examples were matched after filtering.")
    return examples


def split_multisource_examples(
    examples: list[MultiSourceMMSEExample],
    val_ratio: float,
    test_ratio: float,
    seed: int,
) -> dict[str, list[MultiSourceMMSEExample]]:
    grouped: dict[str, list[MultiSourceMMSEExample]] = defaultdict(list)
    for example in examples:
        grouped[example.domain_name].append(example)

    split_sets = {"train": [], "val": [], "test": []}
    for domain_name, domain_examples in grouped.items():
        shuffled = list(domain_examples)
        random.Random(seed + abs(hash(domain_name)) % 10000).shuffle(shuffled)
        total = len(shuffled)
        test_count = max(1, int(total * test_ratio))
        val_count = max(1, int(total * val_ratio))
        train_end = total - val_count - test_count
        if train_end <= 0:
            raise ValueError(f"Domain '{domain_name}' has too few samples for the requested split ratios")
        split_sets["train"].extend(shuffled[:train_end])
        split_sets["val"].extend(shuffled[train_end : train_end + val_count])
        split_sets["test"].extend(shuffled[train_end + val_count :])

    return split_sets


def domain_counts(examples: list[MultiSourceMMSEExample]) -> dict[str, int]:
    counts: dict[str, int] = defaultdict(int)
    for example in examples:
        counts[example.domain_name] += 1
    return dict(sorted(counts.items()))
